# 모듈 불러오기
from discord import File, Embed, app_commands, Interaction, ButtonStyle
from discord.ext import commands
import datetime
import openpyxl
from discord.ui import Button, View

wb = openpyxl.load_workbook('challenge.xlsx')
load_ws = wb['Sheet1']


class challenge(commands.Cog):
    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot

    @app_commands.command(name="메뉴")
    async def challage(self, interaction: Interaction) -> None:
        d = datetime.datetime.now()
        str_now = str(d.strftime("%y-%m-%d\n"))
        # 홀수달 짝수달 구분
        if d.month % 2 == 0:
            # 짝수달 이면
            col = 2
            send_message = ((load_ws.cell(d.day, col).value))
            wb.close()
        else:
            # 홀수달 이면
            col = 1
            send_message = ((load_ws.cell(d.day, col).value))
            wb.close()
        today_embed = Embed(title="오늘의 챌린지",
                            description=str_now + send_message)
        today_file = File(f"./{col}/{d.day}.jpg", filename="image.jpg")
        today_embed.set_image(url="attachment://image.jpg")
        today_embed.set_footer(
            text="루틴표 제공 : 튼튼 한입할게요 \n맵 파일 제공 : 튼튼 한입할게요 \n봇 문의 : 병아리 기현")
        tomorrow = d + datetime.timedelta(days=1)
        strftime_tomorrow = int(tomorrow.strftime("%d"))
        str_tomorrow = str(tomorrow.strftime("%y-%m-%d\n"))
        # 홀수달 짝수달 구분
        if d.month % 2 == 0:
            # 짝수달 이면
            co = 2
            to_challange = ((load_ws.cell(strftime_tomorrow, co).value))
            wb.close()
        else:
            # 홀수달 이면
            co = 1
            to_challange = ((load_ws.cell(strftime_tomorrow, co).value))
            wb.close()
        tomorrow_embed = Embed(title="내일의 챌린지",
                               description=str(str_tomorrow) + to_challange)
        tomorrow_file = File(
            f"./{co}/{strftime_tomorrow}.jpg", filename="image.jpg")
        tomorrow_embed.set_image(url="attachment://image.jpg")
        tomorrow_embed.set_footer(
            text="루틴표 제공 : 튼튼 한입할게요 \n맵 파일 제공 : 튼튼 한입할게요 \n봇 문의 : 병아리 기현")

        today = Button(label="오늘의 챌", style=ButtonStyle.green)
        tomorrow = Button(label="내일의 챌", style=ButtonStyle.primary)
        burning = Button(label="이달의 버닝", style=ButtonStyle.danger)
        tjdanf = Button(label="면류관", style=ButtonStyle.gray)
        tavern = Button(label="tavernofsoul", style=ButtonStyle.link,
                        url="https://ktos.tavernofsoul.com/")

        async def today_callback(interaction: Interaction):
            await interaction.response.send_message(file=today_file, embed=today_embed)

        async def tommorow_callback(interaction: Interaction):
            await interaction.response.send_message(file=tomorrow_file, embed=tomorrow_embed)

        async def burning_callback(interaction: Interaction):
            await interaction.response.send_message(file=File("./img/6b.jpg"))

        async def tjdanf_callback(interaction: Interaction):
            await interaction.response.send_message(file=File("./img/db28631695a79b9e.jpg"))

        today.callback = today_callback
        tomorrow.callback = tommorow_callback
        burning.callback = burning_callback
        tjdanf.callback = tjdanf_callback
        view = View()
        view.add_item(today)
        view.add_item(tomorrow)
        view.add_item(burning)
        view.add_item(tjdanf)
        view.add_item(tavern)
        await interaction.response.send_message("메뉴를 선택해주세요", view=view)


async def setup(bot: commands.Bot) -> None:
    await bot.add_cog(
        challenge(bot)
    )
