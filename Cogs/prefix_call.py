from discord import File, Embed
from discord.ext import commands
import datetime
import openpyxl

wb = openpyxl.load_workbook('challenge.xlsx')
load_ws = wb['Sheet1']


class prefixcall(commands.Cog):
    def __init__(self, bot: commands.Bot) -> None:
        self.bot = bot

    @commands.command(name="챌")
    async def cof(self, ctx):
        d = datetime.datetime.now()
        str_now = str(d.strftime("%y-%m-%d\n"))
        # 홀수달 짝수달 구분
        if d.month % 2 == 0:
            # 짝수달 이면
            col = 2
            send_message = ((load_ws.cell(d.day, col).value))
            wb.close()
        else:
            # 홀수달 이면
            col = 1
            send_message = ((load_ws.cell(d.day, col).value))
            wb.close()
        today_embed = Embed(title="오늘의 챌린지",
                            description=str_now + send_message)
        today_file = File(f"./{col}/{d.day}.jpg", filename="image.jpg")
        today_embed.set_image(url="attachment://image.jpg")
        today_embed.set_footer(
            text="루틴표 제공 : 튼튼 한입할게요 \n맵 파일 제공 : 튼튼 한입할게요 \n봇 문의 : 병아리 기현")
        await ctx.send(file=today_file, embed=today_embed)

    @commands.command(name="내일")
    async def sodlf(self, ctx):
        d = datetime.datetime.now()
        tomorrow = d + datetime.timedelta(days=1)
        strftime_tomorrow = int(tomorrow.strftime("%d"))
        str_tomorrow = str(tomorrow.strftime("%y-%m-%d\n"))
        # 홀수달 짝수달 구분
        if d.month % 2 == 0:
            # 짝수달 이면
            co = 2
            to_challange = ((load_ws.cell(strftime_tomorrow, co).value))
            wb.close()
        else:
            # 홀수달 이면
            co = 1
            to_challange = ((load_ws.cell(strftime_tomorrow, co).value))
            wb.close()
        tomorrow_embed = Embed(title="내일의 챌린지",
                               description=str(str_tomorrow) + to_challange)
        tomorrow_file = File(
            f"./{co}/{strftime_tomorrow}.jpg", filename="image.jpg")
        tomorrow_embed.set_image(url="attachment://image.jpg")
        tomorrow_embed.set_footer(
            text="루틴표 제공 : 튼튼 한입할게요 \n맵 파일 제공 : 튼튼 한입할게요 \n봇 문의 : 병아리 기현")
        await ctx.send(file=tomorrow_file, embed=tomorrow_embed)


async def setup(bot: commands.Bot) -> None:
    await bot.add_cog(
        prefixcall(bot)
    )
